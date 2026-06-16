import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment,
    Border, Side
)
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import pandas as pd
from datetime import datetime

class ExcelHandler:
    """
    Handles all Excel file operations for
    Project Management Dashboard.
    
    Author: Jajitha
    Course: Project Management Tracker - Coursera
    """
    
    def __init__(self, filepath='data/projects.xlsx'):
        self.filepath = filepath
        self.workbook = None
        self.colors = {
            'header': '2196F3',
            'completed': '4CAF50',
            'in_progress': 'FF9800',
            'not_started': 'F44336',
            'white': 'FFFFFF',
            'light_blue': 'E3F2FD'
        }
    
    def create_workbook(self):
        """Create new Excel workbook with sheets"""
        self.workbook = openpyxl.Workbook()
        
        # Create sheets
        self.workbook.active.title = 'Projects'
        self.workbook.create_sheet('Tasks')
        self.workbook.create_sheet('Team')
        self.workbook.create_sheet('Budget')
        self.workbook.create_sheet('Dashboard')
        
        self._setup_projects_sheet()
        self._setup_tasks_sheet()
        self._setup_team_sheet()
        self._setup_budget_sheet()
        self._setup_dashboard_sheet()
        
        self.save()
        print("Excel workbook created successfully!")
    
    def _apply_header_style(self, cell, text):
        """Apply header styling to cell"""
        cell.value = text
        cell.font = Font(
            bold=True,
            color=self.colors['white'],
            size=11
        )
        cell.fill = PatternFill(
            start_color=self.colors['header'],
            end_color=self.colors['header'],
            fill_type='solid'
        )
        cell.alignment = Alignment(
            horizontal='center',
            vertical='center'
        )
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _setup_projects_sheet(self):
        """Setup Projects sheet with headers and sample data"""
        ws = self.workbook['Projects']
        ws.row_dimensions[1].height = 30
        
        headers = [
            'Project ID', 'Project Name', 'Manager',
            'Start Date', 'End Date', 'Status',
            'Progress %', 'Budget', 'Spent', 'Priority'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            self._apply_header_style(cell, header)
            ws.column_dimensions[
                get_column_letter(col)
            ].width = 15
        
        # Sample data
        sample_projects = [
            ['P001', 'GKS Application 2027', 'Jajitha',
             '2026-06-01', '2027-02-28', 'In Progress',
             45, 0, 0, 'High'],
            ['P002', 'TOPIK Preparation', 'Jajitha',
             '2026-06-15', '2026-12-31', 'In Progress',
             10, 0, 0, 'High'],
            ['P003', 'Research Paper NLP', 'Jajitha',
             '2026-07-01', '2026-11-30', 'Not Started',
             0, 0, 0, 'High'],
            ['P004', 'GitHub Portfolio', 'Jajitha',
             '2026-06-14', '2026-06-30', 'In Progress',
             60, 0, 0, 'Medium'],
            ['P005', 'Python Projects', 'Jajitha',
             '2026-06-14', '2026-06-20', 'Completed',
             100, 0, 0, 'Medium'],
        ]
        
        for row_idx, project in enumerate(
            sample_projects, 2
        ):
            for col_idx, value in enumerate(project, 1):
                cell = ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value=value
                )
                cell.alignment = Alignment(
                    horizontal='center'
                )
                
                # Color code by status
                if project[5] == 'Completed':
                    cell.fill = PatternFill(
                        start_color=self.colors['completed'],
                        end_color=self.colors['completed'],
                        fill_type='solid'
                    )
                elif project[5] == 'In Progress':
                    cell.fill = PatternFill(
                        start_color=self.colors['light_blue'],
                        end_color=self.colors['light_blue'],
                        fill_type='solid'
                    )
    
    def _setup_tasks_sheet(self):
        """Setup Tasks sheet"""
        ws = self.workbook['Tasks']
        ws.row_dimensions[1].height = 30
        
        headers = [
            'Task ID', 'Project ID', 'Task Name',
            'Assignee', 'Start Date', 'Due Date',
            'Status', 'Priority', 'Progress %', 'Notes'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            self._apply_header_style(cell, header)
            ws.column_dimensions[
                get_column_letter(col)
            ].width = 15
    
    def _setup_team_sheet(self):
        """Setup Team sheet"""
        ws = self.workbook['Team']
        ws.row_dimensions[1].height = 30
        
        headers = [
            'Member ID', 'Name', 'Role',
            'Email', 'Tasks Assigned',
            'Tasks Completed', 'Availability'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            self._apply_header_style(cell, header)
            ws.column_dimensions[
                get_column_letter(col)
            ].width = 18
    
    def _setup_budget_sheet(self):
        """Setup Budget sheet"""
        ws = self.workbook['Budget']
        ws.row_dimensions[1].height = 30
        
        headers = [
            'Project ID', 'Category', 'Allocated',
            'Spent', 'Remaining', 'Last Updated'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            self._apply_header_style(cell, header)
            ws.column_dimensions[
                get_column_letter(col)
            ].width = 15
    
    def _setup_dashboard_sheet(self):
        """Setup Dashboard summary sheet"""
        ws = self.workbook['Dashboard']
        
        # Title
        ws['A1'] = 'PROJECT MANAGEMENT DASHBOARD'
        ws['A1'].font = Font(
            bold=True,
            size=16,
            color=self.colors['header']
        )
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:F1')
        
        # Summary headers
        summary_headers = [
            ['Metric', 'Value'],
            ['Total Projects', '=COUNTA(Projects!A2:A100)'],
            ['Completed', '=COUNTIF(Projects!F2:F100,"Completed")'],
            ['In Progress', '=COUNTIF(Projects!F2:F100,"In Progress")'],
            ['Not Started', '=COUNTIF(Projects!F2:F100,"Not Started")'],
            ['Last Updated', datetime.now().strftime('%Y-%m-%d')]
        ]
        
        for row_idx, row_data in enumerate(
            summary_headers, 3
        ):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value=value
                )
                if row_idx == 3:
                    self._apply_header_style(cell, value)
                else:
                    cell.alignment = Alignment(
                        horizontal='center'
                    )
                ws.column_dimensions[
                    get_column_letter(col_idx)
                ].width = 20
    
    def read_projects(self):
        """Read projects from Excel file"""
        try:
            df = pd.read_excel(
                self.filepath,
                sheet_name='Projects'
            )
            return df
        except Exception as e:
            print(f"Error reading Excel: {e}")
            return None
    
    def save(self):
        """Save workbook to file"""
        try:
            import os
            os.makedirs(
                os.path.dirname(self.filepath),
                exist_ok=True
            )
            self.workbook.save(self.filepath)
            print(f"Saved to {self.filepath}")
        except Exception as e:
            print(f"Error saving: {e}")


if __name__ == "__main__":
    handler = ExcelHandler()
    handler.create_workbook()
    print("Excel handler initialized successfully!")
